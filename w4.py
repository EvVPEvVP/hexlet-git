import os.path
import random
import string
from openpyxl import load_workbook

def checkthetablestr(y):
    if not os.path.isfile(y):
        return False

    wb = load_workbook(y)
    ws = wb.active
    for k in range(1, 13):
        for i in range(1, 13):
            x = chr(64 + i)
            c = x + str(k)
            # ws[c] = random.randint(1, 100)
            ws[c] = ''.join(random.choice(string.digits) for _ in range(1))
    wb.save(y)
    return True





